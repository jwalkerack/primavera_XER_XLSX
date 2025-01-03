import streamlit as st
from openpyxl import Workbook
import io
def main():
    st.title("Primavera XER -- Excel")
    st.sidebar.header("Choose a Transformation")

    option = st.sidebar.selectbox("Select an option", ("XER to Excel", "Excel to XER"))

    if option == "XER to Excel":
        xer_to_excel()
    elif option == "Excel to XER":
        ### Upload the Excel Button
        excel_to_xer()

import streamlit as st


def format_XER_Uploader(Uploaded):
    xer_file = io.StringIO()
    return xer_file

def Generate_List_Of_XER_Strings(TableIndexDict,FileString):
    StringDict = {}
    for T in TableIndexDict:
        S = TableIndexDict[T]['StartRow']
        E = TableIndexDict[T]['EndRow']
        Rows = []
        for i in range(S,E):
            GETROW = f"{FileString[i]}\n"
            Rows.append(GETROW)
        StringDict[T] = Rows
    return StringDict

def Header_Ender(uploaded_Xer):
    Header = f"{uploaded_Xer[0]}\n"
    Ender = "%E"
    H = [Header,Ender]
    return H


import io
import zipfile
from openpyxl import load_workbook


def get_table_names_from_excel(uploaded_excel):
    """
    Get the list of table names (sheet names) from the uploaded Excel file.

    Args:
        uploaded_excel: The uploaded Excel file object.

    Returns:
        list: A list of table (sheet) names.
    """
    try:
        # Read the file as bytes and create a file-like object
        bytes_data = uploaded_excel.read()  # Read file content into bytes
        file_stream = io.BytesIO(bytes_data)  # Convert bytes into a file-like object

        # Try to load the workbook
        try:
            workbook = load_workbook(file_stream)
        except zipfile.BadZipFile as e:
            # Handle specific zip-related errors (e.g., file is not a valid .xlsx file)
            raise ValueError("Uploaded file is not a valid Excel file. Please upload a valid .xlsx file.")

        # Return the list of sheet names (table names)
        return workbook.sheetnames

    except Exception as e:
        # Catch all other exceptions and raise a more specific error
        raise ValueError(f"Error processing the Excel file: {e}")




def Generate_Dict_Of_Excel_Strings(uploaded_excel,Table_Name):
    excelDict = {}

    # Read the file as bytes (this is where the issue was earlier)
    bytes_data = uploaded_excel.read()  # Read file content into bytes
    # Create a file-like object from the bytes data (this is what we need)
    file_stream = io.BytesIO(bytes_data)  # Convert bytes into a file-like object
    # Now we can load the workbook from the file-like object
    workbook = load_workbook(file_stream)
    # Access the specific sheet
    Data = []
    if Table_Name in workbook.sheetnames:
        sheet = workbook[Table_Name]
        tableName = sheet.title
        RowID = "%T"
        tableString = f"{RowID}\t{tableName}\n"
        Data.append(tableString)
        rowCount = 0
        for row in sheet.iter_rows(values_only=True):
            if rowCount == 0:
                RowID = "%F"
            else:
                RowID = "%T"

            # Create a tab-delimited string for each row
            tab_delimited_string = "\t".join(map(str, row))  # Convert values to strings
            updated_tab_string = f"{RowID}\t{tab_delimited_string}\n"
            Data.append(updated_tab_string)
            rowCount += 1
    excelDict[Table_Name] = Data
    return excelDict

def create_new_xer_content(old_xer_dict, excel_dict, meta):
    """
    Create a new XER file content as a string based on OLDXERDICT, EXCELDICT, and META.

    Args:
        old_xer_dict (dict): Dictionary containing table names and corresponding rows from the old XER.
        excel_dict (dict): Dictionary containing table names and corresponding rows from the Excel input.
        meta (list): A list containing two items:
                     - The first item to be written as the first line of the XER file.
                     - The second item to be written as the last line of the XER file.

    Returns:
        str: The content of the new XER file.
    """
    try:
        # Use StringIO to write the file content in memory
        output = io.StringIO()

        # Add the first meta line as the first line of the file
        output.write(meta[0])

        for table_name, rows in old_xer_dict.items():
            # Determine the source of rows (EXCELDICT takes precedence)
            if table_name in excel_dict:
                source_rows = excel_dict[table_name]
            else:
                source_rows = rows

            # Write the rows to the in-memory buffer
            for row in source_rows:
                output.write(row)  # Ensure each row ends with a newline

        # Add the second meta line as the last line of the file
        output.write(meta[1])

        # Get the in-memory content as a string
        xer_content = output.getvalue()
        output.close()

        return xer_content

    except Exception as e:
        raise RuntimeError(f"An error occurred while creating the XER file: {e}")



import io
def generateNewXERFile(XER_DICT, XER, TABLENAME, TABLEDATA):

    # Create an in-memory file stream to write the new XER data

    xer_file = io.StringIO()
    st.write(TABLENAME)
    st.write(TABLEDATA)
    st.write(XER_DICT)


import io
from openpyxl import load_workbook


def UsingTheExcelLoader(uploaded_file, tableName):
    TableDataCollection = []

    # Read the file as bytes (this is where the issue was earlier)
    bytes_data = uploaded_file.read()  # Read file content into bytes

    # Create a file-like object from the bytes data (this is what we need)
    file_stream = io.BytesIO(bytes_data)  # Convert bytes into a file-like object

    # Now we can load the workbook from the file-like object
    workbook = load_workbook(file_stream)

    # Access the specific sheet
    if tableName in workbook.sheetnames:
        sheet = workbook[tableName]
        tableName = sheet.title
        RowID = "%T"
        tableString = f"{RowID}\t{tableName}\n"
        TableDataCollection.append(tableString)

        rowCount = 0
        for row in sheet.iter_rows(values_only=True):
            if rowCount == 0:
                RowID = "%F"
            else:
                RowID = "%T"

            # Create a tab-delimited string for each row
            tab_delimited_string = "\t".join(map(str, row))  # Convert values to strings
            updated_tab_string = f"{RowID}\t{tab_delimited_string}\n"
            TableDataCollection.append(updated_tab_string)
            rowCount += 1
    else:
        TableDataCollection.append(f"Error: Sheet {tableName} not found.")

    return TableDataCollection

import openpyxl


def load_excel_data(uploaded_file, table_name):
    # Step 1: Read the file content (the file is an UploadedFile object)
    bytes_data = uploaded_file.read()  # Read file content into bytes

    # Step 2: Convert the bytes into a file-like object using io.BytesIO
    file_stream = io.BytesIO(bytes_data)

    # Step 3: Load the workbook from the file-like object
    workbook = openpyxl.load_workbook(file_stream)

    # Step 4: Access the specific sheet
    if table_name in workbook.sheetnames:
        sheet = workbook[table_name]
    else:
        st.error(f"Sheet '{table_name}' not found.")
        return []

    # Step 5: Loop through rows and convert each row (tuple) to a tab-delimited string
    data = []
    for row in sheet.iter_rows(values_only=True):
        # Convert the tuple to a tab-delimited string
        row_string = "\t".join(map(str, row))  # map(str, row) converts each element to string
        data.append(row_string)

    return data

from datetime import datetime
def excel_to_xer():
    st.header("Updates Excel Data into a XER file")

    # File uploader for the XER text file
    xer_fileX = st.file_uploader("Upload Original XER", type=['xer'])
    excel_fileX = st.file_uploader("Upload Original Excel", type=['xlsx'])

    if excel_fileX:
        st.write("Please enter the sheet name:")
        sheet_name = st.text_input("Enter the sheet name:")

    # Allow the user to input the sheet name



    # Trigger the transformation
    if st.button("Transform Excel to Xer"):
        if xer_fileX and excel_fileX:
            try:
                # Attempt to read the XER file
                st.write("Reading XER file...")
                lines = read_file(xer_fileX)  # Reading the XER file
                st.write("Generating List index...")
                gen_list = Generate_List_index(lines)  # Generating list index from XER data
                st.write("Generating Dictionary index...")
                gen_dict = Generate_dict_index(gen_list)  # Creating a dictionary index
                st.write("Generating Table Named Dict...")
                table_dict = Generate_Table_Named_Dict(gen_dict, lines)  # Creating table dictionary
                SD = Generate_List_Of_XER_Strings(table_dict, lines)
                ED = Generate_Dict_Of_Excel_Strings(excel_fileX , sheet_name)
                YourHeader = Header_Ender(lines)
                content = create_new_xer_content(SD, ED,YourHeader)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                file_name = f"updated_XER_{timestamp}.xer"

                st.download_button(
                    label="Download XER File",
                    data=content,
                    file_name=file_name,
                    mime="text/plain"
                )

                # Display the result


            except Exception as e:
                # Show detailed error messages
                st.error(f"An error occurred: {str(e)}")
                st.write(f"Error in function: {e.__traceback__.tb_frame.f_code.co_name}")
                st.write("Please check the traceback above for more details.")

        else:
            st.error("Please upload both files to proceed.")

def xer_to_excel():
    st.header("XER to Excel Transformation")

    # File uploader for the XER text file
    xer_file = st.file_uploader("Upload XER Text File", type=['xer'])

    # Trigger the transformation
    if st.button("Transform XER to Excel"):
        if xer_file:
            try:
                # Read the uploaded XER file and process it
                lines = read_file(xer_file)
                row_ids = Generate_List_index(lines)
                key_dict = Generate_dict_index(row_ids)
                data_dict = generate_data_dict(key_dict, lines)

                # Generate the Excel file in memory
                excel_file = create_dict_to_excel(data_dict)

                # Provide a download button for the generated Excel file
                st.download_button(
                    label="Download Excel File",
                    data=excel_file,
                    file_name="generated_file.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"An error occurred: {e}")
        else:
            st.error("Please upload a file to proceed.")


def read_file(file_obj):
    # Reads lines from the uploaded file-like object
    lines = file_obj.getvalue().decode('latin-1').splitlines()
    return lines

def Generate_List_index(lines):
    RowIds = []
    for line in lines:
        row_element = line[:2]
        RowIds.append(row_element)
    return RowIds

def Generate_dict_index(RowIds):
    DataDict = {}
    A = RowIds
    for Aindex, value in enumerate(A):
        if value == "%T":
            Table = Aindex
            D = []
            SecondPass = A[Aindex + 1:]
            for Bindex, Bvalue in enumerate(SecondPass):
                if Bvalue == "%F":
                    Fields = Aindex + 1
                elif Bvalue == "%R":
                    D.append((Aindex + 1) + Bindex)
                elif Bvalue == "%T" or Bvalue == "%E":
                    break
            DataDict[Table] = {'F': Fields , 'D':D }
    return DataDict
def generate_data_dict(gen_dict,rd_file):
    TableDict = {}
    for item in gen_dict:
        T = []
        f = gen_dict[item]['F']
        d = gen_dict[item]['D']

        table = rd_file[item].split("\t")[-1].strip()
        fields = rd_file[f].split("\t")[1:]
        fields[-1] = fields[-1].strip()
        T.append(fields)
        for v in d:
            row = rd_file[v].split("\t")[1:]
            row[-1] = row[-1].strip()
            T.append(row)
        TableDict[table] = T
    return TableDict

def write_dict_to_excel(data_dict, file_path):
    from openpyxl import Workbook

    # Create a new workbook
    wb = Workbook()

    # Remove the default sheet created in the workbook
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    for table_name, table_data in data_dict.items():
        # Add a new sheet with the table name
        ws = wb.create_sheet(title=table_name)

        # Write the data (header and rows) into the sheet
        for row in table_data:
            ws.append(row)

    # Save the workbook to the specified file path
    wb.save(file_path)
    print(f"Excel file created at: {file_path}")

def create_dict_to_excel(data_dict):
    # Create a new workbook
    wb = Workbook()

    # Remove the default sheet created in the workbook
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    for table_name, table_data in data_dict.items():
        # Add a new sheet with the table name
        ws = wb.create_sheet(title=table_name)

        # Write the data (header and rows) into the sheet
        for row in table_data:
            ws.append(row)

    # Save the workbook to an in-memory file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)  # Move to the beginning of the in-memory file

    return output

def Generate_Table_Named_Dict(gen_dict,rd_file):
    TableDictionary =  {}
    for g in gen_dict:
        GetLastRow = gen_dict[g]['D'][-1]
        GetFirstRow = g
        GetTableName = rd_file[GetFirstRow].split("\t")[-1].strip()
        TableDictionary[GetTableName] = {"StartRow" : GetFirstRow , "EndRow": GetLastRow}
    return TableDictionary


if __name__ == "__main__":
    main()
